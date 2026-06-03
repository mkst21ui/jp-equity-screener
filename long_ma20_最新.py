import numpy as np
import pandas as pd
import yfinance as yf
import sys
import os
import math
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TICKER_MARKET = "^N225"
START_DATE    = "2024-01-01"
END_DATE      = "2026-06-01"

TRAIN_END  = "2025-01-01"
TEST_START = "2024-01-01"

MA_LIST   = [15, 20, 30]
MA_WINDOW = 20
HORIZON   = 20
MIN_SLOPE = 0.0

EV_THRESHOLD = 0.003
RR_THRESHOLD = 1.0

# ===========================================================================
# マクロ設定
# ===========================================================================
MARKET_SCORE_LONG_MIN = 2
OIL_CHANGE_THRESH     =  0.04
USDJPY_WEAK_THRESH    =  0.02
US_RECOVERY_THRESH    =  0.03
RATE_HIKE_MANUAL      = False
MIN_DAILY_LIQUIDITY   = 60_000_000

MA50_WINDOW          = 50
MA_SPREAD_MAX        = 0.10
MA_CONVERGING_BONUS  = 1
MA_DIVERGING_PENALTY = 1

SECTOR_ETF_GROWTH = "1570.T"
SECTOR_ETF_VALUE  = "1306.T"

# ===========================================================================
# [NEW] 売買代金設定
# ===========================================================================
DAIKEN_WINDOW       = 5     # 直近N営業日の平均で判定
DAIKEN_TIER_HIGH    = 5_000_000_000   # 50億円/日 → ランク上位相当
DAIKEN_TIER_MID     = 1_000_000_000   # 10億円/日 → ランク中位相当
# ハードフィルター閾値はMIN_DAILY_LIQUIDITYを継続使用（変更なし）
# 売買代金はスコアに加算せず、情報として記録・表示のみ

# ===========================================================================
ACTIVE_FLAGS_OVERRIDE = {
    "oil_high":        True,
    "yen_weak":        True,
    "rate_hike":       None,
    "us_recovery":     True,
    "ai_semiconductor": True,
}

MACRO_SECTOR_BOOST_LONG = {
    "oil_high": {
        "鉱業":             0,
        "石油・石炭製品":   0,
        "卸売業":           0,
        "海運業":           0,
        "倉庫・運輸関連業": 0,
        "空運業":           1,
        "化学":             0,
        "陸運業":           0,
        "食料品":           0,
    },
    "yen_weak": {
        "輸送用機器": 0,
        "電気機器":   1,
        "精密機器":   0,
        "機械":       0,
        "卸売業":     0,
    },
    "rate_hike": {
        "銀行業":               1,
        "保険業":               1,
        "証券、商品先物取引業": 0,
    },
    "us_recovery": {
        "電気機器":     1,
        "精密機器":     0,
        "情報・通信業": 2,
        "機械":         0,
    },
    "ai_semiconductor": {
        "電気機器":     2,
        "精密機器":     1,
        "情報・通信業": 2,
        "化学":         1,
        "機械":         0,
        "その他製品":   0,
        "金属製品":     1,
    },
}

INDUSTRY_EN_TO_JA = {
    "Agricultural Products":            "水産・農林業",
    "Fishing":                          "水産・農林業",
    "Lumber & Wood Production":         "水産・農林業",
    "Agricultural Operations":          "水産・農林業",
    "Oil & Gas E&P":                    "鉱業",
    "Coal":                             "鉱業",
    "Thermal Coal":                     "鉱業",
    "Coking Coal":                      "鉱業",
    "Other Industrial Metals & Mining": "鉱業",
    "Gold":                             "鉱業",
    "Silver":                           "鉱業",
    "Uranium":                          "鉱業",
    "Oil & Gas Integrated":             "石油・石炭製品",
    "Oil & Gas Refining & Marketing":   "石油・石炭製品",
    "Oil & Gas Midstream":              "石油・石炭製品",
    "Engineering & Construction":       "建設業",
    "Residential Construction":         "建設業",
    "Infrastructure Operations":        "建設業",
    "Packaged Foods":                   "食料品",
    "Beverages—Non-Alcoholic":          "食料品",
    "Beverages - Non-Alcoholic":        "食料品",
    "Beverages—Brewers":                "食料品",
    "Beverages—Wineries & Distilleries":"食料品",
    "Beverages - Wineries & Distilleries":"食料品",
    "Confectioners":                    "食料品",
    "Food Distribution":                "食料品",
    "Textile Manufacturing":            "繊維製品",
    "Apparel Manufacturing":            "繊維製品",
    "Paper & Paper Products":           "パルプ・紙",
    "Packaging & Containers":           "パルプ・紙",
    "Specialty Chemicals":              "化学",
    "Chemicals":                        "化学",
    "Agricultural Inputs":              "化学",
    "Diversified Chemicals":            "化学",
    "Industrial Chemicals":             "化学",
    "Drug Manufacturers—General":               "医薬品",
    "Drug Manufacturers - General":             "医薬品",
    "Drug Manufacturers—Specialty & Generic":   "医薬品",
    "Drug Manufacturers - Specialty & Generic": "医薬品",
    "Biotechnology":                            "医薬品",
    "Healthcare":                               "医薬品",
    "Rubber & Plastics":                "ゴム製品",
    "Tires & Rubber":                   "ゴム製品",
    "Building Materials":               "ガラス・土石製品",
    "Glass & Glass Products":           "ガラス・土石製品",
    "Ceramics":                         "ガラス・土石製品",
    "Steel":                            "鉄鋼",
    "Iron & Steel":                     "鉄鋼",
    "Copper":                           "非鉄金属",
    "Aluminum":                         "非鉄金属",
    "Other Precious Metals & Mining":   "非鉄金属",
    "Metal Fabrication":                "金属製品",
    "Hardware & Accessories":           "金属製品",
    "Farm & Heavy Construction Machinery":   "機械",
    "Specialty Industrial Machinery":        "機械",
    "Industrial Machinery":                  "機械",
    "General Industrial Machinery":          "機械",
    "Tools & Accessories":                   "機械",
    "Pumps & Valves":                        "機械",
    "Pollution & Treatment Controls":        "機械",
    "Electronic Components":                 "電気機器",
    "Consumer Electronics":                  "電気機器",
    "Electrical Equipment & Parts":          "電気機器",
    "Semiconductor Equipment & Materials":   "電気機器",
    "Semiconductors":                        "電気機器",
    "Electronics & Computer Distribution":  "電気機器",
    "Auto Manufacturers":               "輸送用機器",
    "Auto Parts":                       "輸送用機器",
    "Recreational Vehicles":            "輸送用機器",
    "Scientific & Technical Instruments":"精密機器",
    "Medical Devices":                  "精密機器",
    "Medical Instruments & Supplies":   "精密機器",
    "Diagnostics & Research":           "精密機器",
    "Optical Instruments":              "精密機器",
    "Leisure":                          "その他製品",
    "Toys & Hobbies":                   "その他製品",
    "Sporting Goods":                   "その他製品",
    "Electronic Gaming & Multimedia":   "その他製品",
    "Publishing":                       "その他製品",
    "Utilities—Regulated Electric":     "電気・ガス業",
    "Utilities - Regulated Electric":   "電気・ガス業",
    "Utilities—Regulated Gas":          "電気・ガス業",
    "Utilities - Regulated Gas":        "電気・ガス業",
    "Utilities—Diversified":            "電気・ガス業",
    "Utilities - Diversified":          "電気・ガス業",
    "Utilities—Independent Power Producers": "電気・ガス業",
    "Trucking":                         "陸運業",
    "Railroads":                        "陸運業",
    "Integrated Freight & Logistics":   "陸運業",
    "Marine Shipping":                  "海運業",
    "Shipping":                         "海運業",
    "Airlines":                         "空運業",
    "Airports & Air Services":          "空運業",
    "Air Freight & Logistics":          "倉庫・運輸関連業",
    "Courier & Delivery Services":      "倉庫・運輸関連業",
    "Rental & Leasing Services":        "倉庫・運輸関連業",
    "Software—Application":             "情報・通信業",
    "Software - Application":           "情報・通信業",
    "Software—Infrastructure":          "情報・通信業",
    "Software - Infrastructure":        "情報・通信業",
    "Information Technology Services":  "情報・通信業",
    "Telecom Services":                 "情報・通信業",
    "Communication Services":           "情報・通信業",
    "Internet Content & Information":   "情報・通信業",
    "Computer Hardware":                "情報・通信業",
    "Data Storage":                     "情報・通信業",
    "Trading Companies":                "卸売業",
    "Industrial Distribution":          "卸売業",
    "Wholesale":                        "卸売業",
    "Retail":                           "小売業",
    "Grocery Stores":                   "小売業",
    "Department Stores":                "小売業",
    "Discount Stores":                  "小売業",
    "Internet Retail":                  "小売業",
    "Specialty Retail":                 "小売業",
    "Apparel Retail":                   "小売業",
    "Home Improvement Retail":          "小売業",
    "Drug Stores":                      "小売業",
    "Banks—Regional":                   "銀行業",
    "Banks - Regional":                 "銀行業",
    "Banks—Diversified":                "銀行業",
    "Banks - Diversified":              "銀行業",
    "Regional Banks":                   "銀行業",
    "Banks":                            "銀行業",
    "Capital Markets":                  "証券、商品先物取引業",
    "Financial Data & Stock Exchanges": "証券、商品先物取引業",
    "Insurance—Life":                   "保険業",
    "Insurance - Life":                 "保険業",
    "Insurance—Diversified":            "保険業",
    "Insurance - Diversified":          "保険業",
    "Insurance—Property & Casualty":    "保険業",
    "Insurance - Property & Casualty":  "保険業",
    "Insurance—Specialty":              "保険業",
    "Financial Conglomerates":          "その他金融業",
    "Credit Services":                  "その他金融業",
    "Asset Management":                 "その他金融業",
    "Mortgage Finance":                 "その他金融業",
    "Real Estate—General":              "不動産業",
    "Real Estate - General":            "不動産業",
    "Real Estate - Development":        "不動産業",
    "Real Estate - Diversified":        "不動産業",
    "Real Estate Services":             "不動産業",
    "REIT—Diversified":                 "不動産業",
    "REIT—Residential":                 "不動産業",
    "REIT—Office":                      "不動産業",
    "REIT—Industrial":                  "不動産業",
    "REIT—Retail":                      "不動産業",
    "Staffing & Employment Services":   "サービス業",
    "Consulting Services":              "サービス業",
    "Restaurants":                      "サービス業",
    "Hotels & Motels":                  "サービス業",
    "Travel Services":                  "サービス業",
    "Personal Services":                "サービス業",
    "Security & Protection Services":   "サービス業",
    "Education & Training Services":    "サービス業",
    "Advertising Agencies":             "サービス業",
    "Entertainment":                    "サービス業",
    "Broadcasting":                     "サービス業",
}


def normalize_industry(raw: str) -> str:
    if not raw or raw == "未分類":
        return raw or "未分類"
    return INDUSTRY_EN_TO_JA.get(raw, raw)


# ===========================================================================
# Stage1: セクターフィルター（変更なし）
# ===========================================================================
def get_active_macro_flags(macro_state: dict) -> dict:
    active = dict(macro_state)
    for flag, override in ACTIVE_FLAGS_OVERRIDE.items():
        if override is not None:
            active[flag] = override
        elif flag not in active:
            active[flag] = False
    return active


def get_hot_sectors(active_flags: dict) -> dict:
    sector_scores = {}
    for flag, is_active in active_flags.items():
        if not is_active:
            continue
        for sector, boost in MACRO_SECTOR_BOOST_LONG.get(flag, {}).items():
            if boost > 0:
                sector_scores[sector] = sector_scores.get(sector, 0) + boost
    return sector_scores


def stage1_sector_filter(tickers: list, industry_map: dict,
                          active_flags: dict) -> tuple[list, dict]:
    hot_sectors = get_hot_sectors(active_flags)
    if not hot_sectors:
        print("[Stage1] アクティブフラグなし → 全銘柄をそのまま通す")
        return tickers, {t: 0 for t in tickers}

    print(f"[Stage1] HOTセクター: {sorted(hot_sectors.items(), key=lambda x:-x[1])}")
    filtered, boost_map = [], {}
    excluded_log = {}
    for t in tickers:
        ind_raw = industry_map.get(t, "")
        ind_ja  = normalize_industry(ind_raw)
        boost   = hot_sectors.get(ind_ja, 0)
        if boost > 0:
            filtered.append(t)
            boost_map[t] = boost
        else:
            excluded_log[ind_ja] = excluded_log.get(ind_ja, 0) + 1

    print(f"[Stage1] {len(tickers)}銘柄 → {len(filtered)}銘柄（除外業種トップ5: "
          f"{sorted(excluded_log.items(), key=lambda x:-x[1])[:5]}）")
    return filtered, boost_map


# ===========================================================================
# [NEW] 売買代金取得関数
# 設計原則: スコアに加算しない / 情報表示のみ / 過剰適応化ゼロ
#
# 役割:
#   1. 直近DAIKEN_WINDOW日の平均売買代金を計算（静的avg_volumeより正確）
#   2. 流動性ハードフィルターの判定精度を向上（量的改善のみ）
#   3. tier（High/Mid/Low）を判定してExcel表示に使う
#
# 変更しないもの:
#   - EV / RR計算（walk-forward）
#   - macro_boost
#   - Stage1セクターフィルター
# ===========================================================================
def get_daiken_stats(ticker: str) -> dict:
    """
    直近DAIKEN_WINDOW営業日のClose×Volumeから売買代金統計を返す。
    取得失敗時はNaNを返し、呼び出し側でフォールバック処理する。
    """
    default = {
        "daiken_5d_avg": float("nan"),
        "daiken_tier":   "不明",
        "daiken_ok":     True,   # 取得失敗時はフィルターを通す（保守的）
    }
    try:
        df = yf.download(
            ticker,
            period=f"{DAIKEN_WINDOW + 5}d",  # 休場考慮で多めに取得
            auto_adjust=True,
            progress=False,
        )
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)
        if df.empty or "Close" not in df.columns or "Volume" not in df.columns:
            return default

        df = df.dropna(subset=["Close", "Volume"])
        if len(df) < 1:
            return default

        df["daiken"] = df["Close"].squeeze() * df["Volume"].squeeze()
        avg = float(df["daiken"].tail(DAIKEN_WINDOW).mean())

        if avg >= DAIKEN_TIER_HIGH:
            tier = "High (50億+)"
        elif avg >= DAIKEN_TIER_MID:
            tier = "Mid  (10億+)"
        else:
            tier = "Low"

        ok = avg >= MIN_DAILY_LIQUIDITY

        print(f"  [DAIKEN] {ticker}: 直近{DAIKEN_WINDOW}日平均 {avg:,.0f}円 → {tier} / 流動性OK={ok}")

        return {
            "daiken_5d_avg": avg,
            "daiken_tier":   tier,
            "daiken_ok":     ok,
        }

    except Exception as e:
        print(f"  [DAIKEN] {ticker}: 取得失敗 ({e}) → フィルターをスキップ")
        return default


# ===========================================================================
# Excel スタイル定数（変更なし）
# ===========================================================================
COLOR_HEADER_DAY1    = "1F4E79"
COLOR_HEADER_DAY2    = "375623"
COLOR_HEADER_DAY3    = "7B2C2C"
COLOR_HEADER_DAY4    = "4A235A"
COLOR_HEADER_DAY5    = "7E5109"
COLOR_HEADER_MACRO   = "0D4F3C"
COLOR_HEADER_DAIKEN  = "1A5276"   # [NEW] 売買代金セクション用
FONT_BODY = "Arial"


def _cell_border():
    s = Side(border_style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(cell, bg_color, font_color="FFFFFF", bold=True, size=10):
    cell.font      = Font(name=FONT_BODY, bold=bold, color=font_color, size=size)
    cell.fill      = PatternFill("solid", start_color=bg_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _cell_border()

def style_data(cell, bold=False, align="left", num_format=None, bg=None):
    cell.font      = Font(name=FONT_BODY, bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = _cell_border()
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    if num_format:
        cell.number_format = num_format


# ===========================================================================
# 1. 期待値計算（変更なし）
# ===========================================================================
def calc_expected_value(fwd_returns: pd.Series) -> dict:
    if len(fwd_returns) < 5:
        return {"ev": 0.0, "rr_ratio": 0.0, "win_rate": 0.0}
    wins   = fwd_returns[fwd_returns > 0]
    losses = fwd_returns[fwd_returns <= 0]
    win_rate = len(wins) / len(fwd_returns)
    avg_win  = float(wins.mean())   if len(wins)   > 0 else 0.0
    avg_loss = float(losses.mean()) if len(losses) > 0 else 0.0
    ev       = win_rate * avg_win + (1 - win_rate) * avg_loss
    rr_ratio = abs(avg_win / avg_loss) if avg_loss != 0 else 0.0
    return {
        "ev":       round(ev, 5),
        "rr_ratio": round(rr_ratio, 2),
        "win_rate": round(win_rate, 4),
    }


# ===========================================================================
# 2. マクロ状態取得（変更なし）
# ===========================================================================
def _is_accelerating(series: pd.Series, period: int = 5) -> bool:
    if len(series) < period * 2 + 1:
        return False
    chg_recent = (series.iloc[-1]      - series.iloc[-period])   / series.iloc[-period]
    chg_prev   = (series.iloc[-period] - series.iloc[-period*2]) / series.iloc[-period*2]
    return float(chg_recent) > float(chg_prev)


def get_macro_state() -> dict:
    state = {
        "oil_high":    False,
        "yen_weak":    False,
        "rate_hike":   RATE_HIKE_MANUAL,
        "us_recovery": False,
    }
    try:
        oil = yf.download("CL=F", period="2mo", progress=False, auto_adjust=True)
        if not oil.empty:
            if isinstance(oil.columns, pd.MultiIndex):
                oil.columns = oil.columns.get_level_values(0)
            c = oil['Close'].dropna()
            if len(c) >= 22:
                chg_20       = float((c.iloc[-1] - c.iloc[-22]) / c.iloc[-22])
                accelerating = _is_accelerating(c)
                state["oil_high"] = chg_20 > OIL_CHANGE_THRESH and accelerating
                print(f"[MACRO] 原油20日:{chg_20:.2%} 加速中:{accelerating} → oil_high={state['oil_high']}")
    except Exception as e:
        print(f"[MACRO] 原油データ取得失敗: {e}")

    try:
        fx = yf.download("JPY=X", period="2mo", progress=False, auto_adjust=True)
        if not fx.empty:
            if isinstance(fx.columns, pd.MultiIndex):
                fx.columns = fx.columns.get_level_values(0)
            c = fx['Close'].dropna()
            if len(c) >= 22:
                chg_20       = float((c.iloc[-1] - c.iloc[-22]) / c.iloc[-22])
                accelerating = _is_accelerating(c)
                state["yen_weak"] = chg_20 > USDJPY_WEAK_THRESH and accelerating
                print(f"[MACRO] ドル円20日:{chg_20:.2%} 加速中:{accelerating} → yen_weak={state['yen_weak']}")
    except Exception as e:
        print(f"[MACRO] 為替データ取得失敗: {e}")

    try:
        spy = yf.download("SPY", period="2mo", progress=False, auto_adjust=True)
        if not spy.empty:
            if isinstance(spy.columns, pd.MultiIndex):
                spy.columns = spy.columns.get_level_values(0)
            c = spy['Close'].dropna()
            if len(c) >= 6:
                chg_5 = float((c.iloc[-1] - c.iloc[-6]) / c.iloc[-6])
                state["us_recovery"] = chg_5 > US_RECOVERY_THRESH
                print(f"[MACRO] SPY 5日:{chg_5:.2%} → us_recovery={state['us_recovery']}")
    except Exception as e:
        print(f"[MACRO] SPYデータ取得失敗: {e}")

    return state


def calc_macro_boost_long(industry_raw: str, active_flags: dict) -> tuple[int, str]:
    industry = normalize_industry(industry_raw)
    boost, reasons = 0, []
    for flag, sector_map in MACRO_SECTOR_BOOST_LONG.items():
        if not active_flags.get(flag):
            continue
        if industry in sector_map:
            b = sector_map[industry]
            if b > 0:
                boost += b
                reasons.append(f"{flag}×{industry}(+{b})")
    return boost, " | ".join(reasons)


# ===========================================================================
# 3. ロング地合いスコア（変更なし）
# ===========================================================================
def get_market_context_long(symbol: str = "^N225") -> dict:
    try:
        raw = yf.download(symbol, period="1y", progress=False, auto_adjust=True)
        if isinstance(raw.columns, pd.MultiIndex):
            raw.columns = raw.columns.get_level_values(0)
        close = raw['Close'].dropna()
    except Exception as e:
        print(f"[MARKET] 地合いデータ取得失敗: {e}")
        return {"score": 0, "is_favorable": False, "status": "取得失敗"}

    ma20  = close.rolling(20).mean()
    ma50  = close.rolling(50).mean()
    ma200 = close.rolling(200).mean()
    atr14 = close.diff().abs().ewm(alpha=1/14, adjust=False).mean()
    atr_ma20 = atr14.rolling(20).mean()

    latest     = close.iloc[-1]
    base_score = sum([
        latest > float(ma20.iloc[-1]),
        latest > float(ma50.iloc[-1]),
        latest > float(ma200.iloc[-1]),
        float(ma20.iloc[-1]) > float(ma50.iloc[-1]),
        float(atr14.iloc[-1]) < float(atr_ma20.iloc[-1]) * 1.2,
    ])

    sector_score = 0
    try:
        g = yf.download(SECTOR_ETF_GROWTH, period="3mo", progress=False, auto_adjust=True)
        v = yf.download(SECTOR_ETF_VALUE,  period="3mo", progress=False, auto_adjust=True)
        if not g.empty and not v.empty:
            if isinstance(g.columns, pd.MultiIndex): g.columns = g.columns.get_level_values(0)
            if isinstance(v.columns, pd.MultiIndex): v.columns = v.columns.get_level_values(0)
            gc = g['Close'].dropna()
            vc = v['Close'].dropna()
            if len(gc) >= 22 and len(vc) >= 22:
                chg_growth    = float((gc.iloc[-1] - gc.iloc[-22]) / gc.iloc[-22])
                chg_value     = float((vc.iloc[-1] - vc.iloc[-22]) / vc.iloc[-22])
                growth_strong = chg_growth > chg_value
                if growth_strong:
                    sector_score += 1
                print(f"[MARKET] グロース20日:{chg_growth:.2%} / バリュー20日:{chg_value:.2%} "
                      f"→ growth_strong={growth_strong}(+{sector_score})")
    except Exception as e:
        print(f"[MARKET] セクターETF取得失敗: {e}")

    score = base_score + sector_score

    try:
        if len(close) >= MA50_WINDOW + 5:
            ma20_mkt   = close.rolling(MA_WINDOW).mean()
            ma50_mkt   = close.rolling(MA50_WINDOW).mean()
            ma20_now   = float(ma20_mkt.iloc[-1])
            ma50_now   = float(ma50_mkt.iloc[-1])
            ma20_5d    = float(ma20_mkt.iloc[-6])
            ma50_5d    = float(ma50_mkt.iloc[-6])
            spread_now = abs((ma20_now - ma50_now) / ma50_now)
            spread_5d  = abs((ma20_5d  - ma50_5d)  / ma50_5d)
            is_converging = spread_now < spread_5d
            if is_converging:
                score += MA_CONVERGING_BONUS
                print(f"[MARKET] MA収束中（乖離{spread_now:.1%}→{spread_5d:.1%}） → score+{MA_CONVERGING_BONUS}")
            else:
                score -= MA_DIVERGING_PENALTY
                print(f"[MARKET] MA拡散中（乖離{spread_now:.1%}→{spread_5d:.1%}） → score-{MA_DIVERGING_PENALTY}")
        else:
            is_converging = None
    except Exception as e:
        print(f"[MARKET] MA収束チェック失敗: {e}")
        is_converging = None

    is_favorable = score >= MARKET_SCORE_LONG_MIN
    status       = "ロング有利" if score >= 4 else "中立（要注意）" if score >= 2 else "下落地合い"
    print(f"[MARKET] ロング地合いスコア: {score}/6（基本{base_score}+セクター{sector_score}）  {status}")

    return {
        "score":        score,
        "is_favorable": is_favorable,
        "status":       status,
        "ma20":         float(ma20.iloc[-1]),
        "ma50":         float(ma50.iloc[-1]),
        "ma200":        float(ma200.iloc[-1]),
        "latest":       float(latest),
    }


# ===========================================================================
# 4. 出口シグナル計算（変更なし）
# ===========================================================================
def calc_exit_signals_long(entry_price: float, atr: float,
                            take_profit_mult: float = 3.0,
                            trailing_stop_mult: float = 1.5,
                            max_hold_days: int = 50) -> dict:
    if entry_price <= 0 or atr <= 0:
        return {"take_profit": 0.0, "trailing_stop": 0.0, "exit_date_limit": ""}
    take_profit   = round(entry_price + atr * take_profit_mult, 1)
    trailing_stop = round(entry_price - atr * trailing_stop_mult, 1)
    exit_date = datetime.now() + timedelta(days=max_hold_days)
    while exit_date.weekday() >= 5:
        exit_date += timedelta(days=1)
    return {
        "take_profit":     take_profit,
        "trailing_stop":   trailing_stop,
        "exit_date_limit": exit_date.strftime('%Y-%m-%d'),
    }


# ===========================================================================
# 5. 共通ユーティリティ（変更なし）
# ===========================================================================
def get_clean_data(ticker: str, start: str = START_DATE, end: str = END_DATE) -> pd.Series:
    df = yf.download(ticker, start=start, end=end, auto_adjust=True, progress=False)
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)
    if df.empty or "Close" not in df.columns:
        return pd.Series(dtype=float)
    return df["Close"].squeeze().dropna().astype(float)


def get_mae_mfe(close: pd.Series, entries: pd.DatetimeIndex, horizon: int) -> pd.DataFrame:
    rows = []
    idx  = close.index
    for d in entries:
        i    = idx.get_loc(d)
        j    = min(i + horizon, len(close) - 1)
        path = close.iloc[i: j + 1] / close.iloc[i] - 1.0
        rows.append({
            "entry_date": d,
            "entry_px":   float(close.iloc[i]),
            "MAE":        float(path.min()),
            "MFE":        float(path.max()),
            "days":       int(j - i),
        })
    return pd.DataFrame(rows).set_index("entry_date") if rows else pd.DataFrame(
        columns=["entry_px", "MAE", "MFE", "days"])


def calculate_stats(y_vec, x_vec):
    mask = ~(np.isnan(x_vec) | np.isnan(y_vec))
    xv, yv = x_vec[mask], y_vec[mask]
    if len(xv) >= 10:
        beta, alpha = np.polyfit(xv, yv, 1)
        y_hat  = alpha + beta * xv
        ss_res = np.sum((yv - y_hat) ** 2)
        ss_tot = np.sum((yv - np.mean(yv)) ** 2)
        r2     = 1 - (ss_res / ss_tot) if ss_tot != 0 else np.nan
        return alpha, beta, r2
    return np.nan, np.nan, np.nan


def calc_dd(equity: pd.Series) -> pd.Series:
    return (equity / equity.cummax() - 1).astype(float)


def get_max_dd_info(dd_series: pd.Series):
    date = dd_series.idxmin()
    return float(dd_series.loc[date]), date


def _analyze_regression(x, y):
    mask = ~(np.isnan(x) | np.isnan(y))
    xv, yv = x[mask], y[mask]
    if len(xv) < 10:
        return [np.nan, np.nan, np.nan, 0]
    beta, alpha = np.polyfit(xv, yv, 1)
    r2 = 1 - (np.sum((yv - (alpha + beta * xv)) ** 2) / np.sum((yv - np.mean(yv)) ** 2))
    return alpha, beta, r2, len(xv)


def get_ticker_name(ticker: str) -> str:
    try:
        info = yf.Ticker(ticker).info
        name = info.get("longName") or info.get("shortName") or ""
        return name.strip() if name else ticker
    except Exception:
        return ticker


# ===========================================================================
# 7. 銘柄分析
# [変更箇所] get_daiken_stats()の呼び出しと、流動性フィルターの改善
# ===========================================================================
def analyze_ticker(ticker: str, close_m: pd.Series, active_flags: dict,
                   industry_map: dict, ticker_name_map: dict) -> dict | None:
    close_s = get_clean_data(ticker)
    if close_s.empty or len(close_s) < MA_WINDOW + HORIZON:
        print(f"[SKIP] {ticker}: データ不足")
        return None

    # ── [改善] 売買代金取得（静的avg_volumeより精度の高い直近N日平均） ──────
    # 変更点: avg_volume × currentPriceの一時点評価 → 直近5日平均の実績値
    # スコアには影響しない。流動性ハードフィルターの判定精度のみ向上。
    daiken_info = get_daiken_stats(ticker)

    # ── 流動性ハードフィルター ─────────────────────────────────────────────
    # [変更前] yf.Ticker().info から avg_volume × currentPrice を計算
    # [変更後] 直近5日平均の実績売買代金を優先使用、取得失敗時はinfoにフォールバック
    if not pd.isna(daiken_info["daiken_5d_avg"]):
        # 直近実績値で判定
        if not daiken_info["daiken_ok"]:
            print(f"[SKIP] {ticker}: 流動性不足 "
                  f"({daiken_info['daiken_5d_avg']:,.0f}円/日 < {MIN_DAILY_LIQUIDITY:,.0f})")
            return None
    else:
        # フォールバック: 従来のinfo方式
        try:
            info            = yf.Ticker(ticker).info
            avg_volume      = info.get("averageVolume", 0) or 0
            current_price   = (info.get("currentPrice") or
                               info.get("regularMarketPrice") or
                               float(close_s.iloc[-1]))
            daily_liquidity = avg_volume * current_price
            if daily_liquidity < MIN_DAILY_LIQUIDITY:
                print(f"[SKIP] {ticker}: 流動性不足(fallback) "
                      f"({daily_liquidity:,.0f}円/日 < {MIN_DAILY_LIQUIDITY:,.0f})")
                return None
        except Exception as e:
            print(f"[WARN] {ticker}: 流動性チェック失敗 → スキップしない ({e})")

    # ── MA収束フィルター（変更なし） ────────────────────────────────────────
    is_converging: bool | None = None
    try:
        if len(close_s) >= MA50_WINDOW + 5:
            ma20_series = close_s.rolling(MA_WINDOW).mean()
            ma50_series = close_s.rolling(MA50_WINDOW).mean()
            ma20_now    = float(ma20_series.iloc[-1])
            ma50_now    = float(ma50_series.iloc[-1])
            ma20_5d     = float(ma20_series.iloc[-6])
            ma50_5d     = float(ma50_series.iloc[-6])
            spread_now  = abs((ma20_now - ma50_now) / ma50_now)
            spread_5d   = abs((ma20_5d  - ma50_5d)  / ma50_5d)
            if spread_now > MA_SPREAD_MAX:
                print(f"[SKIP] {ticker}: MA拡散 ({spread_now:.1%} > {MA_SPREAD_MAX:.0%})")
                return None
            is_converging = spread_now < spread_5d
            print(f"  {ticker}: MA乖離={spread_now:.1%} "
                  f"({'収束中↓' if is_converging else '拡散中↑'})")
    except Exception as e:
        print(f"[WARN] {ticker}: MA収束チェック失敗 ({e})")

    # ── 以下: EV / walk-forward計算（変更なし） ────────────────────────────
    _close_m = close_m.reindex(close_s.index).ffill()
    ret_s    = close_s.pct_change().fillna(0)
    ret_m    = _close_m.pct_change().fillna(0)

    train_s          = close_s[:TRAIN_END]
    ma20_train       = train_s.rolling(MA_WINDOW).mean()
    entry_cond_train = (train_s > ma20_train) & (ma20_train.diff() > MIN_SLOPE)
    fwd_ret_train    = (train_s.shift(-HORIZON) / train_s - 1.0)
    valid_train      = fwd_ret_train[entry_cond_train].dropna()
    ev_train         = calc_expected_value(valid_train)

    test_s = close_s[TEST_START:]
    if test_s.empty:
        print(f"[WARN] {ticker}: テスト期間データなし。全期間で評価。")
        test_s = close_s

    ma20_full          = close_s.rolling(MA_WINDOW).mean()
    slope_full         = ma20_full.diff()
    entry_cond_test    = (
        (test_s > ma20_full.reindex(test_s.index)) &
        (slope_full.reindex(test_s.index) > MIN_SLOPE)
    ).fillna(False)
    entry_trigger_test = entry_cond_test & (~entry_cond_test.shift(1, fill_value=False))
    entry_dates        = test_s.index[entry_trigger_test]

    if len(entry_dates) == 0:
        print(f"[SKIP] {ticker}: テスト期間にエントリーシグナルなし")
        return None

    fwd_ret_s = (close_s.shift(-HORIZON) / close_s - 1.0).loc[entry_dates]
    fwd_ret_m = (_close_m.shift(-HORIZON) / _close_m - 1.0).loc[entry_dates]
    win_rate  = float((fwd_ret_s > 0).mean()) if len(fwd_ret_s) > 0 else 0.0
    ev_result = calc_expected_value(fwd_ret_s.dropna())

    print(f"  {ticker}: 訓練EV={ev_train['ev']:.3%} / テストEV={ev_result['ev']:.3%} "
          f"/ 勝率={win_rate:.2%} / RR={ev_result['rr_ratio']:.2f}")

    ma20    = close_s.rolling(MA_WINDOW).mean()
    pos     = (close_s > ma20).shift(1).fillna(0).astype(int)

    events          = get_mae_mfe(close_s, entry_dates, HORIZON)
    alpha, beta, r2 = calculate_stats(fwd_ret_s.values, fwd_ret_m.values)

    strat_ret = pos * ret_s
    eq_strat  = (1 + strat_ret).cumprod()
    eq_bh     = (1 + ret_s).cumprod()
    eq_mkt    = (1 + ret_m).cumprod()

    dd_strat = calc_dd(eq_strat)
    dd_bh    = calc_dd(eq_bh)
    dd_mkt   = calc_dd(eq_mkt)

    max_dd_s,  date_s  = get_max_dd_info(dd_strat)
    max_dd_bh, date_bh = get_max_dd_info(dd_bh)
    max_dd_m,  date_m  = get_max_dd_info(dd_mkt)

    is_active = pos == 1
    res_exp   = _analyze_regression(ret_m.values, strat_ret.values)
    res_act   = _analyze_regression(ret_m[is_active].values, ret_s[is_active].values)

    trough_date         = dd_strat.idxmin()
    trough_val          = dd_strat.min()
    peak_date           = eq_strat[:trough_date].idxmax()
    recovery_candidates = eq_strat[trough_date:]
    peak_val            = eq_strat[peak_date]
    recovery_date       = (
        recovery_candidates[recovery_candidates >= peak_val].index.min()
        if (recovery_candidates >= peak_val).any() else None
    )

    robustness_rows = []
    for n in MA_LIST:
        ma_n          = close_s.rolling(n).mean()
        slope_n       = ma_n.diff()
        pos_n         = (close_s > ma_n).shift(1).fillna(0).astype(int)
        eq_n          = (1 + pos_n * ret_s).cumprod()
        dd_n          = calc_dd(eq_n)
        is_entry_n    = (close_s > ma_n) & (slope_n > MIN_SLOPE)
        entry_dates_n = close_s.index[is_entry_n & (~is_entry_n.shift(1, fill_value=False))]
        events_n      = get_mae_mfe(close_s, entry_dates_n, HORIZON)
        robustness_rows.append({
            "MA":         f"MA{n}",
            "Entries":    len(entry_dates_n),
            "Return":     float(eq_n.iloc[-1]),
            "MaxDD":      float(dd_n.min()),
            "MaxDD_Date": dd_n.idxmin().date(),
            "MAE_5pct":   float(events_n["MAE"].quantile(0.05)) if not events_n.empty else np.nan,
            "MFE_Med":    float(events_n["MFE"].median())        if not events_n.empty else np.nan,
        })

    industry_raw = industry_map.get(ticker, "")
    macro_boost, macro_reason = calc_macro_boost_long(industry_raw, active_flags)

    latest_price = float(close_s.iloc[-1])
    latest_atr   = float(close_s.diff().abs().ewm(alpha=1/14, adjust=False).mean().iloc[-1])
    exit_signals = calc_exit_signals_long(latest_price, latest_atr)

    ticker_name = ticker_name_map.get(ticker, ticker)
    industry_ja = normalize_industry(industry_raw)

    return {
        "ticker":          ticker,
        "name":            ticker_name,
        "industry":        industry_ja,
        "industry_raw":    industry_raw,
        "win_rate":        win_rate,
        "ev":              ev_result["ev"],
        "rr_ratio":        ev_result["rr_ratio"],
        "ev_train":        ev_train["ev"],
        "entries":         len(entry_dates),
        "stock_fwd":       float(fwd_ret_s.mean()),
        "market_fwd":      float(fwd_ret_m.mean()),
        "mae_worst":       float(events["MAE"].min())    if not events.empty else np.nan,
        "mfe_median":      float(events["MFE"].median()) if not events.empty else np.nan,
        "alpha":           float(alpha) if pd.notna(alpha) else np.nan,
        "beta":            float(beta)  if pd.notna(beta)  else np.nan,
        "r2":              float(r2)    if pd.notna(r2)    else np.nan,
        "dd_strat":        max_dd_s,
        "dd_strat_date":   date_s.date(),
        "dd_bh":           max_dd_bh,
        "dd_bh_date":      date_bh.date(),
        "dd_mkt":          max_dd_m,
        "dd_mkt_date":     date_m.date(),
        "exp_beta":        float(res_exp[1]) if pd.notna(res_exp[1]) else np.nan,
        "exp_r2":          float(res_exp[2]) if pd.notna(res_exp[2]) else np.nan,
        "act_beta":        float(res_act[1]) if pd.notna(res_act[1]) else np.nan,
        "act_r2":          float(res_act[2]) if pd.notna(res_act[2]) else np.nan,
        "dd4_depth":       float(trough_val),
        "dd4_peak":        peak_date.date(),
        "dd4_trough":      trough_date.date(),
        "dd4_days":        (trough_date - peak_date).days,
        "dd4_recovery":    str(recovery_date.date()) if recovery_date else "未回復",
        "robustness":      robustness_rows,
        "macro_boost":     macro_boost,
        "macro_reason":    macro_reason,
        "is_converging":   is_converging,
        "latest_price":    latest_price,
        "latest_atr":      latest_atr,
        "take_profit":     exit_signals["take_profit"],
        "trailing_stop":   exit_signals["trailing_stop"],
        "exit_date_limit": exit_signals["exit_date_limit"],
        # [NEW] 売買代金情報（スコアには不使用 / 表示のみ）
        "daiken_5d_avg":   daiken_info["daiken_5d_avg"],
        "daiken_tier":     daiken_info["daiken_tier"],
    }


# ===========================================================================
# 8. Excel書き出し
# ===========================================================================
def write_summary_sheet(wb: Workbook, results: list, active_flags: dict, market_ctx: dict):
    ws       = wb.create_sheet("Summary")
    ws.title = "Summary"

    ws.merge_cells("A1:L1")
    ws["A1"] = "ロングスキャン結果サマリー（v5 マクロ連動・期待値ベース）"
    ws["A1"].font      = Font(name=FONT_BODY, bold=True, size=14, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A2"] = (f"分析期間: {START_DATE} ～ {END_DATE}  |  "
                f"EVフィルタ: {EV_THRESHOLD:.1%}以上 / RR: {RR_THRESHOLD:.1f}倍以上  |  "
                f"地合い: {market_ctx.get('status','')}")
    ws["A2"].font = Font(name=FONT_BODY, italic=True, size=9, color="555555")
    ws.merge_cells("A2:L2")

    ws["A3"] = "アクティブフラグ: " + "  /  ".join(
        f"{k}={'ON' if v else 'off'}" for k, v in active_flags.items()
    )
    ws["A3"].font = Font(name=FONT_BODY, bold=True, size=9,
                         color="0D4F3C" if any(active_flags.values()) else "888888")
    ws.merge_cells("A3:L3")

    hot_sectors = get_hot_sectors(active_flags)
    ws["A4"] = "HOTセクター: " + "  ".join(
        f"{s}(+{sc})" for s, sc in sorted(hot_sectors.items(), key=lambda x: -x[1])
    )
    ws["A4"].font = Font(name=FONT_BODY, bold=True, size=9, color="7E5109")
    ws.merge_cells("A4:L4")

    # [変更] 売買代金Tierカラムを追加（K列）
    headers = ["Ticker", "銘柄名", "業種", "Win Rate", "EV", "RR比",
               "Macro Boost", "Macro根拠", "Entries", "Take Profit", "Trailing Stop",
               "売買代金Tier"]
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(row=6, column=col, value=h), "1F4E79")

    for r, res in enumerate(results, 7):
        passed = res.get("ev", 0) >= EV_THRESHOLD and res.get("rr_ratio", 0) >= RR_THRESHOLD
        tier   = res.get("daiken_tier", "不明")
        vals = [
            res["ticker"],
            res.get("name", ""),
            res.get("industry", ""),
            res["win_rate"],
            res.get("ev", 0),
            res.get("rr_ratio", 0),
            res.get("macro_boost", 0),
            res.get("macro_reason", ""),
            res["entries"],
            res.get("take_profit", ""),
            res.get("trailing_stop", ""),
            tier,
        ]
        fmts   = [None, None, None, "0.00%", "0.000%", "0.00", "0", None, "0", "0.0", "0.0", None]
        row_bg = ("E8F5E9" if (passed and res.get("macro_boost", 0) > 0)
                  else "F0FFF4" if passed
                  else "FFEBEE")
        for col, (v, fmt) in enumerate(zip(vals, fmts), 1):
            style_data(ws.cell(row=r, column=col, value=v), align="center", num_format=fmt, bg=row_bg)

    for i, w in enumerate([14, 22, 18, 10, 10, 8, 12, 35, 8, 12, 12, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A7"


def write_ticker_sheet(wb: Workbook, res: dict):
    ticker = res["ticker"]
    ws     = wb.create_sheet(ticker)
    row    = 1

    def write_section_header(label, color, row):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        c = ws.cell(row=row, column=1, value=label)
        style_header(c, color, size=11)
        ws.row_dimensions[row].height = 22
        return row + 1

    def write_kv(key, value, row, num_format=None, value_bold=False):
        ck = ws.cell(row=row, column=1, value=key)
        style_data(ck, bold=True, bg="F2F2F2")
        cv = ws.cell(row=row, column=2, value=value)
        style_data(cv, bold=value_bold, align="right", num_format=num_format)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        return row + 1

    passed = res.get("ev", 0) >= EV_THRESHOLD and res.get("rr_ratio", 0) >= RR_THRESHOLD
    ws.merge_cells("A1:H1")
    ws["A1"] = (
        f"{ticker}  {res.get('name', '')}  [{res.get('industry','')}]  "
        f"EV: {res.get('ev',0):.3%}  RR: {res.get('rr_ratio',0):.2f}  "
        f"Macro Boost: +{res.get('macro_boost',0)}"
    )
    ws["A1"].font      = Font(name=FONT_BODY, bold=True, size=13, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", start_color="1F4E79" if passed else "7B2C2C")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    row = 2

    # ── [NEW] 売買代金セクション ───────────────────────────────────────────
    # Macro セクションの直前に配置。スコアへの影響はゼロ。
    # 表示目的: 人間が「本当に資金が入っているか」を確認するための参照情報。
    row = write_section_header("売買代金  流動性・市場関心度", COLOR_HEADER_DAIKEN, row)
    daiken_avg = res.get("daiken_5d_avg", float("nan"))
    daiken_avg_display = f"{daiken_avg:,.0f}円" if not pd.isna(daiken_avg) else "取得不可"
    row = write_kv("直近5日平均売買代金", daiken_avg_display,         row)
    row = write_kv("売買代金ティア",      res.get("daiken_tier", "不明"), row, value_bold=True)
    row = write_kv(
        "参考",
        "High=50億+/日(外国人機関レベル)  Mid=10億+/日  Low=それ以下",
        row,
    )
    row += 1

    # ── Macro セクション（変更なし） ──────────────────────────────────────
    row = write_section_header("Macro  マクロ連動・出口シグナル", COLOR_HEADER_MACRO, row)
    row = write_kv("業種（日本語）",         res.get("industry", "不明"),     row)
    row = write_kv("業種（原文）",           res.get("industry_raw", ""),      row)
    row = write_kv("マクロブースト",         res.get("macro_boost", 0),        row, "0", value_bold=True)
    row = write_kv("マクロ根拠",             res.get("macro_reason", "なし"),  row)
    is_conv    = res.get("is_converging")
    conv_label = "収束中↓" if is_conv is True else "拡散中↑" if is_conv is False else "判定不能"
    row = write_kv("MA収束状態",             conv_label,                        row)
    row = write_kv("現在値",                 res.get("latest_price", 0),       row, "0.0")
    row = write_kv("ATR（14日）",            res.get("latest_atr", 0),         row, "0.00")
    row = write_kv("利確目標（ATR×3）",      res.get("take_profit", 0),        row, "0.0", value_bold=True)
    row = write_kv("損切ライン（ATR×1.5）", res.get("trailing_stop", 0),      row, "0.0", value_bold=True)
    row = write_kv("時間切れ決済日",         res.get("exit_date_limit", ""),    row)
    row += 1

    # ── Day1-5（変更なし） ────────────────────────────────────────────────
    row = write_section_header("Day1  エントリー統計", COLOR_HEADER_DAY1, row)
    row = write_kv("Sample Period",           f"{START_DATE}  →  {END_DATE}",  row)
    row = write_kv("Walk-forward Train End",  TRAIN_END,                        row)
    row = write_kv("Walk-forward Test Start", TEST_START,                       row)
    row = write_kv("Entries Found (test)",    res["entries"],                   row, "0")
    row = write_kv("Stock  Mean Fwd Ret",     res["stock_fwd"],                 row, "0.00%")
    row = write_kv("Market Mean Fwd Ret",     res["market_fwd"],                row, "0.00%")
    row = write_kv("Win Rate",                res["win_rate"],                  row, "0.00%")
    row = write_kv("EV (期待値)",             res.get("ev", 0),                 row, "0.000%", value_bold=True)
    row = write_kv("RR比",                    res.get("rr_ratio", 0),           row, "0.00", value_bold=True)
    row = write_kv("EV 訓練期",               res.get("ev_train", 0),           row, "0.000%")
    row = write_kv("MAE (Max Adverse Worst)", res["mae_worst"],                 row, "0.00%")
    row = write_kv("MFE (Max Fav Median)",    res["mfe_median"],                row, "0.00%")
    row = write_kv("Alpha",                   res["alpha"],                     row, "0.0000")
    row = write_kv("Beta",                    res["beta"],                      row, "0.00")
    row = write_kv("R-squared",               res["r2"],                        row, "0.00")
    row += 1

    row = write_section_header("Day2  ドローダウン", COLOR_HEADER_DAY2, row)
    for col, h in enumerate(["Metric", "Max DD", "Date"], 1):
        style_header(ws.cell(row=row, column=col, value=h), "375623")
    row += 1
    for metric, dd, date in [
        ("Strategy",      res["dd_strat"],  res["dd_strat_date"]),
        ("Buy & Hold",    res["dd_bh"],     res["dd_bh_date"]),
        ("Market (N225)", res["dd_mkt"],    res["dd_mkt_date"]),
    ]:
        style_data(ws.cell(row=row, column=1, value=metric), bold=True, bg="F2F2F2")
        style_data(ws.cell(row=row, column=2, value=dd), align="right", num_format="0.00%",
                   bg="FFE0E0" if dd < -0.20 else "FFF3CD" if dd < -0.10 else "E8F5E9")
        style_data(ws.cell(row=row, column=3, value=str(date)), align="center")
        row += 1
    row += 1

    row = write_section_header("Day3  ベータ分析", COLOR_HEADER_DAY3, row)
    row = write_kv("Exposure Beta", res["exp_beta"], row, "0.00")
    row = write_kv("Exposure R²",   res["exp_r2"],   row, "0.00")
    row = write_kv("Active Beta",   res["act_beta"], row, "0.00")
    row = write_kv("Active R²",     res["act_r2"],   row, "0.00")
    row += 1

    row = write_section_header("Day4  MaxDD 解剖", COLOR_HEADER_DAY4, row)
    row = write_kv("下落の深さ",      res["dd4_depth"],       row, "0.00%")
    row = write_kv("ピーク日",        str(res["dd4_peak"]),   row)
    row = write_kv("トラフ日",        str(res["dd4_trough"]), row)
    row = write_kv("下落期間 (days)", res["dd4_days"],        row, "0")
    row = write_kv("回復状態",        res["dd4_recovery"],    row, value_bold=True)
    row += 1

    row = write_section_header("Day5  ロバスト性", COLOR_HEADER_DAY5, row)
    day5_headers = ["MA", "Entries", "Return (x)", "Max DD", "Max DD Date", "MAE 5%ile", "MFE Median"]
    day5_fmts    = [None,  "0",       "0.00",       "0.00%",   None,          "0.00%",     "0.00%"]
    for col, h in enumerate(day5_headers, 1):
        style_header(ws.cell(row=row, column=col, value=h), "7E5109")
    row += 1
    for rb in res["robustness"]:
        vals = [rb["MA"], rb["Entries"], rb["Return"], rb["MaxDD"],
                str(rb["MaxDD_Date"]), rb["MAE_5pct"], rb["MFE_Med"]]
        for col, (v, fmt) in enumerate(zip(vals, day5_fmts), 1):
            style_data(ws.cell(row=row, column=col, value=v), align="center", num_format=fmt,
                       bg="F9F9F9" if row % 2 == 0 else "FFFFFF")
        row += 1

    for i, w in enumerate([28, 14, 14, 14, 18, 18, 18, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


# ===========================================================================
# 9. メイン（変更なし）
# ===========================================================================
if __name__ == "__main__":
    excel_path  = sys.argv[1] if len(sys.argv) >= 2 else "longv2.xlsx"
    output_path = "output_long_v5_daiken.xlsx"

    if not os.path.exists(excel_path):
        print(f"[ERROR] {excel_path} が見つかりません")
        sys.exit(1)

    raw_df = pd.read_excel(excel_path, dtype=str)
    raw_df.columns = ["銘柄コード", "業種"] + list(raw_df.columns[2:])
    raw_df = raw_df.dropna(subset=["銘柄コード"])
    raw_df["銘柄コード"] = raw_df["銘柄コード"].str.strip()
    raw_df["業種"]       = raw_df["業種"].fillna("未分類").str.strip()
    raw_df["銘柄コード"] = raw_df["銘柄コード"].apply(
        lambda t: f"{t}.T" if str(t).isdigit() else t
    )

    tickers      = raw_df["銘柄コード"].tolist()
    industry_map = dict(zip(raw_df["銘柄コード"], raw_df["業種"]))

    print(f"[INFO] {len(tickers)} 銘柄を読み込みました")

    print("\n=== マクロ状態を取得中 ===")
    macro_state  = get_macro_state()
    active_flags = get_active_macro_flags(macro_state)

    print("\n[v5] アクティブフラグ（自動+手動合成）:")
    for k, v in active_flags.items():
        override = ACTIVE_FLAGS_OVERRIDE.get(k)
        source   = "手動" if override is not None else "自動"
        print(f"  {k:20s}: {'ON' if v else 'off'}  ({source})")

    print("\n=== ロング地合い評価 ===")
    market_ctx = get_market_context_long()
    if not market_ctx["is_favorable"]:
        print(
            f"\n地合いスコア {market_ctx['score']}/6 — ロング非推奨局面です。\n"
            f"   含み損がある場合は静観 or ヘッジを優先してください。\n"
            f"   処理は続行しますが、macro_boostが0の銘柄は除外を検討してください。"
        )

    print(f"\n=== [Stage1] セクターフィルター（{len(tickers)}銘柄 → 絞り込み）===")
    filtered_tickers, _ = stage1_sector_filter(tickers, industry_map, active_flags)

    if not filtered_tickers:
        print("[ERROR] Stage1通過銘柄なし。ACTIVE_FLAGS_OVERRIDEを確認してください。")
        sys.exit(1)

    print(f"\n=== 銘柄名を取得中（{len(filtered_tickers)}銘柄）===")
    ticker_name_map = {}
    for t in filtered_tickers:
        ticker_name_map[t] = get_ticker_name(t)
        print(f"  {t}: {ticker_name_map[t]}")

    print("\n=== 市場データ（N225）取得中 ===")
    close_m_global = get_clean_data(TICKER_MARKET)

    print(f"\n=== [Stage2] 銘柄分析（{len(filtered_tickers)}銘柄）===")
    all_results, skipped = [], []
    for ticker in filtered_tickers:
        print(f"[分析中] {ticker} ...")
        try:
            res = analyze_ticker(ticker, close_m_global, active_flags,
                                  industry_map, ticker_name_map)
            if res is None:
                skipped.append(ticker)
            else:
                all_results.append(res)
        except Exception as e:
            print(f"[ERROR] {ticker}: {e}")
            skipped.append(ticker)

    wb = Workbook()
    wb.remove(wb.active)

    passed_results = [r for r in all_results
                      if r.get("ev", 0) >= EV_THRESHOLD and r.get("rr_ratio", 0) >= RR_THRESHOLD]
    failed_results = [r for r in all_results
                      if not (r.get("ev", 0) >= EV_THRESHOLD and r.get("rr_ratio", 0) >= RR_THRESHOLD)]

    write_summary_sheet(wb, all_results, active_flags, market_ctx)
    for res in sorted(all_results, key=lambda x: (-x.get("ev", 0), -x.get("macro_boost", 0))):
        write_ticker_sheet(wb, res)

    wb.save(output_path)

    print("\n" + "=" * 60)
    print(f"[Stage1] 入力: {len(tickers)}銘柄 → 通過: {len(filtered_tickers)}銘柄")
    print(f"[Stage2] EV・RR通過: {len(passed_results)} 銘柄 / 未通過: {len(failed_results)} 銘柄")
    print()
    for r in sorted(passed_results, key=lambda x: (-x.get("ev", 0), -x.get("macro_boost", 0))):
        boost_str   = f" [Macro+{r['macro_boost']}]" if r.get("macro_boost", 0) > 0 else ""
        daiken_str  = f" [{r.get('daiken_tier','不明')}]"
        print(f"   {r['ticker']:<12}  {r.get('name',''):<30}  "
              f"EV:{r.get('ev',0):.3%}  RR:{r.get('rr_ratio',0):.2f}  "
              f"業種:{r.get('industry','')}{boost_str}{daiken_str}")
    if skipped:
        print(f"\nスキップ: {', '.join(skipped)}")
    print(f"\n出力完了: {output_path}")
    print(f"\n=== アクティブフラグサマリー ===")
    for k, v in active_flags.items():
        print(f"  {k:20s}: {'ON' if v else 'off'}")
    print(f"  地合いスコア        : {market_ctx['score']}/6  {market_ctx.get('status','')}")